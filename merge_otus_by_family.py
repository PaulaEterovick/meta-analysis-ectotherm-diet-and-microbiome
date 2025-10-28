#!/usr/bin/env python3
"""
Adapted script to merge OTUs by family in genus_level_correlated.xlsx
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os
import shutil

def main():
    # File path
    excel_file = "/Users/paulaeterovick/Dokumente/ProjektDFG2024/metanalysis/phyloseq_merged/genus_level_correlated.xlsx"
    
    # Check if file exists
    if not os.path.exists(excel_file):
        print(f"Error: File {excel_file} not found!")
        return
    
    print("Reading Excel file...")
    
    # Read both sheets
    try:
        # Load workbook to see all sheet names
        wb = load_workbook(excel_file, read_only=True)
        print(f"Available sheets: {wb.sheetnames}")
        
        # Read the sheets
        otu_correlated = pd.read_excel(excel_file, sheet_name='original_otu_correlated')
        taxonomy_correlated = pd.read_excel(excel_file, sheet_name='original_taxonomy_correlated')
        
        print("\n=== OTU CORRELATED SHEET ===")
        print(f"Shape: {otu_correlated.shape}")
        print(f"Columns: {list(otu_correlated.columns[:10])}" + ("..." if len(otu_correlated.columns) > 10 else ""))
        print("\nFirst few rows:")
        print(otu_correlated.head(3))
        
        print("\n=== TAXONOMY CORRELATED SHEET ===")
        print(f"Shape: {taxonomy_correlated.shape}")
        print(f"Columns: {list(taxonomy_correlated.columns)}")
        print("\nFirst few rows:")
        print(taxonomy_correlated.head(3))
        
        # Check that both sheets have the same OTUs
        otu_counts_otus = set(otu_correlated['otu'].dropna())
        taxonomy_otus = set(taxonomy_correlated['otu'].dropna())
        
        print(f"\nOTUs in count data: {len(otu_counts_otus)}")
        print(f"OTUs in taxonomy data: {len(taxonomy_otus)}")
        
        # Find common OTUs
        common_otus = otu_counts_otus.intersection(taxonomy_otus)
        print(f"Common OTUs: {len(common_otus)}")
        
        if len(common_otus) != len(otu_counts_otus) or len(common_otus) != len(taxonomy_otus):
            print("Warning: Not all OTUs match between sheets. Using only common OTUs.")
            
            # Filter both sheets to common OTUs
            otu_correlated = otu_correlated[otu_correlated['otu'].isin(common_otus)]
            taxonomy_correlated = taxonomy_correlated[taxonomy_correlated['otu'].isin(common_otus)]
        
        # Analyze family distribution
        print("\n=== FAMILY ANALYSIS ===")
        family_counts = taxonomy_correlated['Family'].value_counts()
        print(f"Total unique families: {len(family_counts)}")
        print(f"Families with multiple OTUs: {sum(family_counts > 1)}")
        print(f"Maximum OTUs per family: {family_counts.max()}")
        
        # Show top families with most OTUs
        print("\nTop 10 families with most OTUs:")
        print(family_counts.head(10))
        
        # Merge OTU data with taxonomy to get family information
        print("\n=== MERGING DATA BY FAMILY ===")
        merged_data = otu_correlated.merge(taxonomy_correlated, on='otu', how='inner')
        print(f"Merged data shape: {merged_data.shape}")
        
        # Identify sample columns (all columns except taxonomic ones)
        sample_columns = [col for col in otu_correlated.columns if col != 'otu']
        taxonomy_columns = ['Phylum', 'Class', 'Order', 'Family', 'Genus', 'Species']
        
        print(f"Sample columns: {len(sample_columns)}")
        print(f"Sample columns (first 10): {sample_columns[:10]}")
        
        # Group by family and aggregate
        print("\nGrouping OTUs by family and summing counts...")
        
        # For each family, we'll:
        # 1. Sum all the sample counts
        # 2. Keep the taxonomic information (taking first occurrence for higher levels)
        
        agg_dict = {}
        
        # Sum sample columns
        for col in sample_columns:
            agg_dict[col] = 'sum'
        
        # For taxonomic columns (except Family), take the first occurrence
        # This assumes OTUs within the same family have consistent higher-level taxonomy
        for col in taxonomy_columns:
            if col != 'Family':
                agg_dict[col] = 'first'
        
        # Group by family and aggregate
        family_aggregated = merged_data.groupby('Family').agg(agg_dict).reset_index()
        
        print(f"\nAggregated data shape: {family_aggregated.shape}")
        print(f"Reduced from {len(otu_correlated)} OTUs to {len(family_aggregated)} families")
        
        # Reorder columns to match original structure
        column_order = ['Family'] + taxonomy_columns[:-1] + sample_columns  # Family first, then other taxonomy, then samples
        column_order = [col for col in column_order if col in family_aggregated.columns and col != 'Family']
        final_columns = ['Family'] + column_order
        
        family_aggregated = family_aggregated[final_columns]
        
        print("\nFinal aggregated data:")
        print(family_aggregated.head())
        
        # Create taxonomy sheet for family level
        print("\n=== CREATING FAMILY TAXONOMY SHEET ===")
        
        # Only include taxonomy columns that actually exist in the aggregated data
        available_tax_cols = [col for col in taxonomy_columns if col in family_aggregated.columns]
        family_taxonomy = family_aggregated[available_tax_cols].copy()
        
        # Create abundance sheet (family + sample counts)
        print("\n=== CREATING FAMILY ABUNDANCE SHEET ===")
        
        available_sample_cols = [col for col in sample_columns if col in family_aggregated.columns]
        family_abundance = family_aggregated[['Family'] + available_sample_cols].copy()
        
        print(f"Family taxonomy shape: {family_taxonomy.shape}")
        print(f"Family abundance shape: {family_abundance.shape}")
        
        # Save results
        print("\n=== SAVING RESULTS ===")
        
        # Create backup
        backup_file = excel_file.replace('.xlsx', '_backup.xlsx')
        print(f"Creating backup: {backup_file}")
        shutil.copy2(excel_file, backup_file)
        
        # Create new file with family-level data
        output_file = excel_file.replace('.xlsx', '_family_merged.xlsx')
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write family-level aggregated sheets
            family_abundance.to_excel(writer, sheet_name='Family_abundance', index=False)
            family_taxonomy.to_excel(writer, sheet_name='family_tax', index=False)
            
            # Also write the original sheets for reference
            otu_correlated.to_excel(writer, sheet_name='original_otu_correlated', index=False)
            taxonomy_correlated.to_excel(writer, sheet_name='original_taxonomy_correlated', index=False)
        
        print(f"\nResults saved to: {output_file}")
        print(f"Backup created: {backup_file}")
        
        # Summary statistics
        print("\n=== SUMMARY ===")
        print(f"Original OTUs: {len(otu_correlated)}")
        print(f"Final families: {len(family_aggregated)}")
        print(f"Reduction ratio: {len(otu_correlated)/len(family_aggregated):.2f}x")
        print(f"Sample columns preserved: {len(sample_columns)}")
        
        # Show some examples of merged families
        multi_otu_family = family_counts[family_counts > 1]
        if len(multi_otu_family) > 0:
            print(f"\nExamples of families with multiple OTUs merged:")
            for family, count in multi_otu_family.head(5).items():
                original_otus = taxonomy_correlated[taxonomy_correlated['Family'] == family]['otu'].tolist()
                print(f"  {family}: {count} OTUs merged ({original_otus})")
        
        print("\nFamily-level merging completed successfully!")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
